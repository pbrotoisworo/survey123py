import json
from pathlib import Path
import shutil
import yaml
from dataclasses import dataclass

import pandas as pd
import openpyxl
from pyxform.xls2json import parse_file_to_json
from pyxform.errors import PyXFormError

@dataclass
class Sheets:
    """
    Class to store constant value sheet names for Survey123

    They are in the order:
    - Survey
    - Choices
    - Settings
    - Version
    - Question types
    - Appearances
    - Field types
    - Reference
    - Reserved
    """
    survey: str = "survey"
    choices: str = "choices"
    settings: str = "settings"
    version: str = "Version"
    question_types: str = "Question types"
    appearances: str = "Appearances"
    field_types: str = "Field types"
    reference: str = "Reference"
    reserved: str = "Reserved"


class FormData:
    """
    Class to handle form data.
    """

    def __init__(self, version: str):
        self._module_dir = Path(__file__).parent
        self.sheets = {
            Sheets.survey: None,
            Sheets.choices: None,
            Sheets.settings: None,
            Sheets.version: None,
            Sheets.question_types: None,
            Sheets.appearances: None,
            Sheets.field_types: None,
            Sheets.reference: None,
            Sheets.reserved: None
        }
        self._template_paths = {
            "3.22": {
                "survey": self._module_dir / "template" / "template_3.22.xlsx",
                "columns": self._module_dir / "template" / "template_3.22_columns.json",
            },
        }
        self.yaml_data = None
        self.form_version = None
        self._load_template(version)
        

    def _load_template(self, version: str):
        """
        Load Survey123 template according to the version
        and store it in the sheets property to be filled in later.

        Parameters
        ----------
        version : str
            Template version being used for Survey123
        """

        
        if version not in self._template_paths.keys():
            raise ValueError(f"Version {version} not supported. Supported versions are: {list(self._template_paths.keys())}")

        for sheet_name in self.sheets.keys():
            try:
                self.sheets[sheet_name] = pd.read_excel(self._template_paths[version][Sheets.survey], sheet_name=sheet_name)
            except ValueError as e:
                print(f"Error loading sheet {sheet_name}: {e}")
        self.form_version = self.sheets[Sheets.version].iloc[1]["Unnamed: 1"]
    
    def load_yaml(self, path: str):
        """
        Load YAML file containing survey data and convert it to a DataFrame
        to be stored in the sheets property.
        
        Once loaded. The available sheets are:
        - survey
        - choices
        - settings
        - version
        - Question types
        - Appearances
        - Field types
        - Reference
        - Reserved
        
        You can access the different sheets using the code sample below:
        ```python
        survey = FormData("3.22")
        survey.load_yaml("path/to/survey.yaml")
        survey.sheets["survey"]  # Access the survey sheet
        survey.sheets["choices"]  # Access the choices sheet
        ```

        Parameters
        ----------
        path : str
            Path to the survey data file.
        """
        # Load YAML file
        with open(path, 'r') as file:
            survey_data = yaml.safe_load(file)
            self.yaml_data = survey_data

        # Validate
        with open(self._template_paths[self.form_version]["columns"]) as f:
            template_cols = json.load(f)

        if Sheets.survey in survey_data.keys():
            self.sheets[Sheets.survey] = self._load_yaml_survey_sheet(survey_data, template_cols)
        if Sheets.choices in survey_data.keys():
            self.sheets[Sheets.choices] = self._load_yaml_choices_sheet(survey_data, template_cols)
        if Sheets.settings in survey_data.keys():
            self.sheets[Sheets.settings] = self._load_yaml_settings_sheet(survey_data)
                

    def _load_yaml_settings_sheet(self, survey_data):
        # Due to the way settings are structured, it is not in a list.
        # We need to convert it to a list so that it can be loaded into a DataFrame.
        if isinstance(survey_data[Sheets.settings], dict):
            survey_data[Sheets.settings] = [survey_data[Sheets.settings]]
        return pd.DataFrame(survey_data[Sheets.settings])
    
    def _load_yaml_choices_sheet(self, survey_data, template_cols):
        choices_data_processed = []
        for field in survey_data[Sheets.choices]:
            choices_data_processed.append(field)
        df_input = pd.DataFrame(choices_data_processed)
        df_input["name"] = df_input["name"].replace({True: "yes", False: "no"})
        df_input["label"] = df_input["label"].replace({True: "yes", False: "no"})
        df = pd.DataFrame(columns=template_cols[Sheets.choices])
        return pd.concat([df, df_input], axis=0, ignore_index=True)
    
    def _load_yaml_survey_sheet(self, survey_data, template_cols):
        survey_data_processed = []
        for _, field in enumerate(survey_data[Sheets.survey]):
            
            # Process groups and add the begin/end group statements automatically
            if field["type"] == "group":
                survey_data_processed.append({"type": "begin group", "name": field["name"], "label": field["label"]})
                for child in field["children"]:
                    survey_data_processed.append(child)
                survey_data_processed.append({"type": "end group"})
                continue

            # Process repeats and add the begin/end group statements automatically
            if field["type"] == "repeat":
                survey_data_processed.append({"type": "begin repeat", "name": field["name"], "label": field["label"]})
                for child in field["children"]:
                    survey_data_processed.append(child)
                survey_data_processed.append({"type": "end repeat"})
                continue
            survey_data_processed.append(field)
    
        df_input = pd.DataFrame(survey_data_processed)
        if "readonly" not in df_input.columns:
            df_input["readonly"] = ""
        if "required" not in df_input.columns:
            df_input["required"] = ""
        df_input["readonly"] = df_input["readonly"].map({True: "yes"})
        df_input["required"] = df_input["required"].map({True: "yes"})
        
        
        df = pd.DataFrame(columns=template_cols[Sheets.survey])
        return pd.concat([df, df_input], axis=0, ignore_index=True)

    def save_survey(self, outpath: str):
        """
        Save the survey data to the specified path.
        """
        # Copy template file to the output path and load it using openpyxl
        # to preserve formatting and styles
        shutil.copy(self._template_paths[self.form_version][Sheets.survey], outpath)
        wb = openpyxl.load_workbook(outpath)

        with open(self._template_paths[self.form_version]["columns"]) as f:
            template_cols = json.load(f)
        
        # Iterate through the sheets and write the data to the corresponding sheets
        target_sheets = [Sheets.survey, Sheets.choices, Sheets.settings]
        for sheet_name in list(self.yaml_data.keys()):
            sheet_data = self.sheets.get(sheet_name, None)
            if sheet_data is None or sheet_name not in target_sheets:
                continue
            
            ws = wb[sheet_name]
            excel_col = None
            excel_row = None

            if sheet_name == Sheets.settings:
                input_data = sheet_data.iloc[0].to_dict()
                for col in template_cols["settings"]:
                    excel_col = template_cols["settings"][col]
                    excel_row = 2
                    target_cell = f"{excel_col}{excel_row}"
                    ws[target_cell].value = input_data.get(col, None)

            if sheet_name == Sheets.survey or sheet_name == Sheets.choices:

                # Start at row 3 to give some space between column titles and data
                excel_row = 3

                for _, row in sheet_data.iterrows():
                    input_data = row.to_dict()
                    if sheet_name == Sheets.choices:
                        for col in template_cols["choices"]:
                            excel_col = template_cols["choices"][col]
                            target_cell = f"{excel_col}{excel_row}"
                            cell_data = input_data.get(col, None)
                            # if pd.isna(cell_data) is False:
                            #     print("Writing to cell", target_cell, ":", cell_data)
                            ws[target_cell].value = cell_data
                        excel_row += 1
                    
                    if sheet_name == Sheets.survey:
                        for col in template_cols["survey"]:
                            excel_col = template_cols["survey"][col]
                            target_cell = f"{excel_col}{excel_row}"
                            cell_data = input_data.get(col, None)
                            ws[target_cell].value = cell_data
                        excel_row += 1

        wb.save(outpath)
        
        # Validate output file is valid
        is_valid, error = self._validate_survey(outpath)
        if is_valid is False:
            print(f"Warning: Output file is not valid.")
            print(error)
        

    def _validate_survey(self, path) -> bool:
        """
        Validate the .xlsx file is valid according to ODK standard.
        """
        try:
            survey = parse_file_to_json(path)
            return True, None
        except PyXFormError as e:
            return False, e

if __name__ == "__main__":
    pass
